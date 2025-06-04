from fastapi import FastAPI, Request, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from datetime import datetime
from typing import Optional, List
from sqlalchemy.orm import Session
import psutil
import subprocess
from database import get_db
from models import ActivityLog, Employee
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_active_window_info():
    try:
        # Using xdotool to get window information on Linux
        active_window_id = subprocess.check_output(['xdotool', 'getactivewindow']).decode().strip()
        window_title = subprocess.check_output(['xdotool', 'getwindowname', active_window_id]).decode().strip()
        
        # Get process info using psutil
        pid = int(subprocess.check_output(['xdotool', 'getwindowpid', active_window_id]).decode().strip())
        process = psutil.Process(pid)
        app_name = process.name()
        
        return {
            "app": app_name,
            "title": window_title,
            "pid": pid
        }
    except Exception as e:
        return {
            "app": "Unknown",
            "title": "Unknown",
            "pid": None,
            "error": str(e)
        }

@app.get("/current-activity")
async def get_current_activity(db: Session = Depends(get_db)):
    activity = get_active_window_info()
    
    if activity["app"] != "Unknown":
        log = ActivityLog(
            employee_id=1,  # Временно хардкодим ID сотрудника
            window_title=activity["title"],
            app_name=activity["app"],
            start_time=datetime.utcnow()
        )
        db.add(log)
        db.commit()
    
    return activity

@app.post("/stop")
async def stop_day(request: Request, db: Session = Depends(get_db)):
    data = await request.json()

    try:
        start_time = datetime.fromisoformat(data["start_time"])
        end_time = datetime.fromisoformat(data["end_time"])
        
        # Закрываем последнюю активность
        last_activity = db.query(ActivityLog)\
            .filter(ActivityLog.employee_id == 1)\
            .filter(ActivityLog.end_time.is_(None))\
            .first()
            
        if last_activity:
            last_activity.end_time = end_time
            db.commit()
            
    except Exception as e:
        return {"error": f"Неверный формат времени: {e}"}

    duration = (end_time - start_time).total_seconds()
    return {
        "message": "Рабочий день завершен",
        "start": start_time.isoformat(),
        "end": end_time.isoformat(),
        "duration_seconds": duration,
        "duration_human": f"{int(duration // 3600)} ч {int((duration % 3600) // 60)} мин"
    }

@app.get("/report")
async def generate_report(
    start_date: str,
    end_date: str,
    department: Optional[str] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)
    
    # Базовый запрос
    query = db.query(
        ActivityLog,
        Employee
    ).join(
        Employee,
        ActivityLog.employee_id == Employee.id
    ).filter(
        ActivityLog.start_time >= start,
        ActivityLog.start_time <= end
    )
    
    # Применяем фильтры
    if department:
        query = query.filter(Employee.department == department)
    if employee_id:
        query = query.filter(Employee.id == employee_id)
    
    activities = query.all()
    
    # Обработка результатов
    report_data = {}
    for activity, employee in activities:
        if employee.id not in report_data:
            report_data[employee.id] = {
                "employeeId": employee.id,
                "employeeName": employee.name,
                "department": employee.department,
                "totalTime": 0,
                "activeTime": 0,
                "apps": {}
            }
        
        # Расчет времени
        duration = (activity.end_time - activity.start_time).total_seconds() / 60
        report_data[employee.id]["totalTime"] += duration
        
        # Учет времени по приложениям
        if activity.app_name not in report_data[employee.id]["apps"]:
            report_data[employee.id]["apps"][activity.app_name] = 0
        report_data[employee.id]["apps"][activity.app_name] += duration
    
    # Формируем итоговый отчет
    items = list(report_data.values())
    total_time = sum(item["totalTime"] for item in items)
    
    return {
        "items": items,
        "totalDepartmentTime": total_time,
        "averageEmployeeTime": total_time / len(items) if items else 0
    }

@app.get("/export-excel")
async def export_excel(
    start_date: str,
    end_date: str,
    department: Optional[str] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    # Получаем данные отчета
    report_data = await generate_report(start_date, end_date, department, employee_id, db)
    
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по активности"
    
    # Стили для заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Заголовки
    headers = ["Сотрудник", "Отдел", "Общее время (ч)", "Активное время (ч)", "Приложения"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Данные
    row = 2
    for item in report_data["items"]:
        ws.cell(row=row, column=1, value=item["employeeName"])
        ws.cell(row=row, column=2, value=item["department"])
        ws.cell(row=row, column=3, value=round(item["totalTime"] / 60, 2))
        ws.cell(row=row, column=4, value=round(item["activeTime"] / 60, 2))
        
        # Форматируем список приложений
        apps_str = "\n".join([f"{app}: {round(time/60, 2)}ч" for app, time in item["apps"].items()])
        ws.cell(row=row, column=5, value=apps_str)
        row += 1
    
    # Итоги
    row += 1
    ws.cell(row=row, column=1, value="Итого по отделу:")
    ws.cell(row=row, column=3, value=round(report_data["totalDepartmentTime"] / 60, 2))
    row += 1
    ws.cell(row=row, column=1, value="Среднее на сотрудника:")
    ws.cell(row=row, column=3, value=round(report_data["averageEmployeeTime"] / 60, 2))
    
    # Автоматическая ширина колонок
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Сохраняем в буфер
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Формируем имя файла
    filename = f"report_{start_date}_{end_date}.xlsx"
    
    # Возвращаем файл
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/employees")
async def get_employees(db: Session = Depends(get_db)):
    employees = db.query(Employee).all()
    return [{"id": emp.id, "name": emp.name, "department": emp.department} for emp in employees]